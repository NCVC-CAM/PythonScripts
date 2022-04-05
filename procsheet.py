import sys
import os
import re
import pandas as pd
import openpyxl as xl
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side

################################################################################
## procssheet.py Ver1.00
################################################################################
## 使い方
## python procsheet.py sample.ncd [sample.xlsx]
## 最後の出力ファイル名は省略可．省略すると 入力ファイル.xlsx で出力されます．
################################################################################

## 引数処理
inFile=sys.argv[1]
outFile=os.path.splitext(sys.argv[2 if len(sys.argv)>2 else 1])[0]+'.xlsx'
#print(outFile)

## 無視する行とワード処理する正規表現
reg_ignore = re.compile(r'^%|\(.*?\)|[\r\n]$')
reg_progno = re.compile(r'O\d+')
reg_word   = re.compile(r'[A-Z/]-?\d*\.?\d*')

## プロセスシートの列をDataFrameで作成
col = ['/','N','G','X','Y','Z','R/I','J','K','F','S','T','M','H/D','L','P','Q']
df = pd.DataFrame(index=[], columns=col)

prog=''     ## O番号
## 1行（ブロック）ずつ処理
with open(inFile) as f:
    for line in f:
        ## 先頭の%とカッコ内のコメントと行末の改行は''に置換
        line = reg_ignore.sub('', line)
        if len(line)==0:
            continue
        ## プログラム番号
        wordlist = reg_progno.findall(line)
        for word in wordlist:
            prog += word + ' '
        ## ブロックをワードごとに分割
        dic = {}
        wordlist = reg_word.findall(line)
        for word in wordlist:
            ## ワードの先頭1文字をキーに各列に文字列を追加
            w = word[:1]
            if w=='R' or w=='I':
                w = 'R/I'
            elif w=='H' or w=='D':
                w = 'H/D'
            elif w=='O':    ## O番号は列として登録しない
                continue
            dic[w] = (dic.get(w) or '') + word
        ## DataFrameに1行追加
        df = df.append(dic, ignore_index=True)

#print(df)
df.to_excel(outFile, index=False, startrow=4)   ## 4行空けて出力

## ヘッダーセルへの代入
wb = xl.load_workbook(outFile)
ws = wb.active
ws['A1'] = '名称'
ws['A3'] = 'プログラム番号'
ws['A4'] = prog
ws['E1'] = '注釈'
ws['N1'] = '日付'
ws['N3'] = '作成者'

## 列タイトルに網掛け
for row in ws['A5:Q5']:
    for cell in row:
        cell.fill = PatternFill(patternType='mediumGray')

## 罫線の設定
side = Side(style='thin', color='000000')

## ヘッダーの罫線を引く
## 冗長な書き方ですが，やむなし？？
top   = Border(top=side)
left  = Border(left=side)
right = Border(right=side)
for row in ws['A1:Q1']:
    for cell in row:
        cell.border = top
for row in ws['A3:D3']:
    for cell in row:
        cell.border = top
for row in ws['N3:Q3']:
    for cell in row:
        cell.border = top
for row in ws['A1:A4']:
    for cell in row:
        cell.border += left
for row in ws['E1:E4']:
    for cell in row:
        cell.border += left
for row in ws['N1:N4']:
    for cell in row:
        cell.border += left
for row in ws['Q1:Q4']:
    for cell in row:
        cell.border += right

## データの罫線を引く
border = Border(top=side, bottom=side, left=side, right=side)
for row in ws.iter_rows(min_row=5):     ## 5行目から
    for cell in row:
        ws[cell.coordinate].border = border

## 印刷設定
## 横方向，列合わせ，タイトル行
ws.print_title_rows = '5:5'
ps = ws.page_setup
ps.orientation = 'landscape'
ps.paperSize = ws.PAPERSIZE_A4
ps.fitToWidth = 1
ps.fitToHeight = 0
ws.sheet_properties.pageSetUpPr.fitToPage = True

wb.save(outFile)
